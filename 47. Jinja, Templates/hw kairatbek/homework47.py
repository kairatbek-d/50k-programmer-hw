from flask import Flask, render_template
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

app = Flask(__name__)

@app.route('/')
def homepage():
    excel = Workbook()
    page = excel.active
    page["A1"] = "Items"
    page["B1"] = "Quantity"
    page["A2"] = "Coins"
    page["B2"] = 23
    page["A3"] = "Chairs"
    page["B3"] = 33
    page["A4"] = "pencils"
    page["B4"] = 43
    excel.save('testing.xlsx')

    return render_template('index.html')

@app.route('/readExcel')
def readpage():
    values = {}
    page = load_workbook('testing.xlsx')
    excel = page.active
    cells = excel['A2': 'B4']
    for c1, c2 in cells:
        values[c1.value] = c2.value
    return render_template('readExcel.html', values=values, headings=[excel["A1"].value, excel["B1"].value])

if __name__ == '__main__':
    app.debug = True
    app.run()