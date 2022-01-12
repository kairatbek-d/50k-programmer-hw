from flask import Flask, render_template, request
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

app = Flask(__name__)

@app.route('/')
def homepage():
    excel = Workbook()
    page = excel.active
    page["A1"] = "Items"
    page["B1"] = "Quantity"
    page["A2"] = "Coins"
    page["B2"] = 23
    page["A3"] = "Chairs"
    page["B3"] = 33
    page["A4"] = "pencils"
    page["B4"] = 43
    excel.save('testing.xlsx')

    return render_template('index.html')

@app.route('/readExcel')
def readpage():
    values = {}
    page = load_workbook('testing.xlsx')
    excel = page.active
    cells = excel['A2': 'B15']
    headings=[
        ["__|", excel["A1"].column_letter, excel["B1"].column_letter],
        [ excel["A1"].row, excel["A1"].value, excel["B1"].value]
        ]
    for c1, c2 in cells:
        if c1.value != None or c2.value != None:
            values[c1.value] = [c2.row, c2.value]
    return render_template('readExcel.html', values=values, headings=headings)

@app.route('/add/', methods=["post"])
def add():
    number = request.form["number"]
    item = request.form["item"]
    quantity = request.form["quantity"]
    
    excel = load_workbook('testing.xlsx')
    page = excel.active
    page[f"A{number}"] = item
    page[f"B{number}"] = quantity
    excel.save('testing.xlsx')

    return """
        <h1>Успешно добавлен</h1>
        <a href="/readExcel">Click to see excel values</a>
    """

if __name__ == '__main__':
    app.debug = True
    app.run()